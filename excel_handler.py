import pandas as pd
import openpyxl

def import_excel(excel_path):
    if not excel_path.is_file():
        raise FileNotFoundError(f'Arquivo {excel_path} não encontrado.')

    xls = pd.ExcelFile(excel_path)
    dfs = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}

    return dfs

def save_excel(data, excel_path, new_excel_path, sheet_name=None):
    """
    Atualiza uma aba específica de uma planilha Excel com os dados de um DataFrame do pandas.
    
    :param data: DataFrame do pandas contendo os dados a serem escritos na planilha.
    :param excel_path: Caminho para o arquivo Excel que será atualizado.
    :param new_excel_path: Caminho para o novo arquivo Excel que será salvo.
    :param sheet_name: Nome da aba que será atualizada com os dados do DataFrame.
    """
    wb = openpyxl.load_workbook(excel_path)

    if isinstance(data, pd.DataFrame):
        if sheet_name is None:
            raise ValueError('O nome da aba deve ser especificado para salvar o DataFrame em um novo arquivo.')
        update_sheet(data, wb, sheet_name, new_excel_path)
    elif isinstance(data, dict):
        for sheet, df in data.items():
            update_sheet(df, wb, sheet, new_excel_path)
    else:
        raise TypeError('O tipo do dado deve ser um DataFrame ou um dicionário de DataFrames.')
    
    wb.save(new_excel_path)
    print(f'Arquivo salvo com sucesso.')

def update_sheet(data, wb, sheet_name, new_excel_path):
    """
    Atualiza ou cria uma aba na workbook com os dados de um DataFrame do pandas, 
    utilizando openpyxl para manter as formatações e configurações existentes.

    :param data: DataFrame do pandas com os dados a serem escritos na aba.
    :param wb: Workbook do openpyxl que está sendo atualizada.
    :param sheet_name: Nome da aba a ser atualizada ou criada.
    :param new_excel_path: Caminho para o novo arquivo Excel que será salvo.
    """
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    for row_index, row in data.iterrows():
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)

    wb.save(new_excel_path)
